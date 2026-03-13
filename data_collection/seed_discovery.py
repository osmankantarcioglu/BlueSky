"""
Load seed users from the Excel file into the TrackedUser table.

For each row that has a bsky_handle, the script:
  1. Resolves the handle → DID via the AT Protocol API
  2. Infers stance from party name
  3. Upserts a TrackedUser record

Usage (run from project root):
    python data_collection/seed_discovery.py
"""
import os
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import time
import openpyxl
from atproto import Client
from database.models import db, TrackedUser, create_tables
from config.settings import BSKY_HANDLE, BSKY_APP_PASSWORD

# Excel file path (relative to project root)
EXCEL_PATH = "bsky_manual_minimal.xlsx"

# How many API calls to make per second (stay well under rate limits)
RATE_LIMIT_DELAY = 0.5  # seconds between requests

# Party name → stance mapping
ALLIANCE_PARTIES = {
    "adalet ve kalkınma partisi", "ak parti", "akp",
    "milliyetçi hareket partisi", "mhp",
}
OPPOSITION_PARTIES = {
    "cumhuriyet halk partisi", "chp",
    "halkların demokratik partisi", "hdp",
    "dem parti", "dem",
    "iyi parti", "i̇yi parti",
    "gelecek partisi",
    "deva partisi", "deva",
    "zafer partisi",
    "türkiye işçi partisi", "tip",
}


def party_to_stance(party: str) -> str:
    """Map a party name string to 'alliance', 'opposition', or 'unknown'."""
    if not party:
        return "unknown"
    p = party.strip().lower()
    if p in ALLIANCE_PARTIES:
        return "alliance"
    if p in OPPOSITION_PARTIES:
        return "opposition"
    return "unknown"


def load_excel(path: str) -> list[dict]:
    """Read the Excel file and return rows that have a bsky_handle."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Read header row
    headers = [cell.value for cell in ws[1]]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        handle = record.get("bsky_handle")
        if handle and str(handle).strip():
            rows.append(record)

    print(f"Excel loaded: {ws.max_row - 1} total rows, {len(rows)} have a bsky_handle")
    return rows


def resolve_and_save(client: Client, rows: list[dict]) -> dict:
    """Resolve each handle to a DID and upsert into TrackedUser."""
    stats = {"saved": 0, "skipped": 0, "errors": 0}

    for i, row in enumerate(rows, 1):
        handle = str(row["bsky_handle"]).strip().lower()
        # Remove leading '@' if present
        handle = handle.lstrip("@")

        party = str(row.get("party") or "").strip()
        name  = f"{row.get('name', '')} {row.get('surname', '')}".strip()
        stance = party_to_stance(party)

        try:
            result = client.com.atproto.identity.resolve_handle({"handle": handle})
            did = result.did

            TrackedUser.get_or_create(
                did=did,
                defaults={
                    "handle":       handle,
                    "display_name": name,
                    "party":        party,
                    "stance":       stance,
                    "domain":       "politics",
                    "source":       "excel",
                    "is_active":    True,
                },
            )
            stats["saved"] += 1

            if i % 20 == 0:
                print(f"  Progress: {i}/{len(rows)} | saved={stats['saved']} errors={stats['errors']}")

        except Exception as e:
            stats["errors"] += 1
            print(f"  [WARN] Could not resolve '{handle}': {e}")

        time.sleep(RATE_LIMIT_DELAY)

    return stats


def main():
    # Make sure tables exist
    create_tables()

    if db.is_closed():
        db.connect()

    # Load Excel
    rows = load_excel(EXCEL_PATH)
    if not rows:
        print("No rows with bsky_handle found. Exiting.")
        return

    # Log in to Bluesky (needed to call resolve_handle)
    if not BSKY_HANDLE or not BSKY_APP_PASSWORD:
        print("ERROR: Set BSKY_HANDLE and BSKY_APP_PASSWORD in your .env file.")
        return

    client = Client()
    client.login(BSKY_HANDLE, BSKY_APP_PASSWORD)
    print(f"Logged in as {BSKY_HANDLE}")

    # Resolve and save
    print(f"Resolving {len(rows)} handles (this will take ~{len(rows) * RATE_LIMIT_DELAY:.0f}s)...")
    stats = resolve_and_save(client, rows)

    db.close()

    print("\n=== Done ===")
    print(f"  Saved : {stats['saved']}")
    print(f"  Errors: {stats['errors']}")
    print(f"  Total handles attempted: {stats['saved'] + stats['errors']}")


if __name__ == "__main__":
    main()
